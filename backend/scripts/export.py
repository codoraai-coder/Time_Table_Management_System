"""
Export timetable to Excel (V2 format).
Usage:
    python scripts/export.py [version_id]
"""
import sys
import os
from dotenv import load_dotenv
# Path setup
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, "../.."))
if os.path.basename(project_root) == "backend":
    project_root = os.path.abspath(os.path.join(project_root, ".."))
sys.path.insert(0, project_root)
sys.path.insert(0, os.path.join(project_root, "backend"))
load_dotenv()
from app.core.database import SessionLocal
from app.services.excel_exporter import ExcelExporterServiceV2
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from app.models import Course

def main():
    db = SessionLocal()
    try:
        version_id = int(sys.argv[1]) if len(sys.argv) > 1 else 1
        print(f">> Exporting Timetable Version {version_id} to Excel...\n")
        exporter = ExcelExporterServiceV2(db)
        out_path = f"timetable_v{version_id}.xlsx"
        ok = exporter.export_timetable(
            version_id=version_id,
            output_path=out_path
        )
        if not ok:
            print("✗ Exporter failed")
            return

        # Post-process workbook to merge consecutive LAB slots (two-hour labs)
        print("[*] Post-processing Excel to merge LAB slots...")
        wb = load_workbook(out_path)

        # Build course type map
        course_map = {c.code: (c.type or '').upper() for c in db.query(Course).all()}
        thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for ws in wb.worksheets:
            # Only process section sheets (start with 'Section ')
            if not ws.title.startswith('Section '):
                continue

            max_col = ws.max_column
            # find last timetable row: first row where A cell == 'Faculty Assignment'
            last_row = None
            for r in range(3, ws.max_row + 1):
                a_val = ws.cell(row=r, column=1).value
                if a_val and str(a_val).strip().startswith('Faculty Assignment'):
                    last_row = r - 2
                    break
            if not last_row:
                last_row = 3 + 5 - 1  # fallback: 5 days

            # Process rows 3..(last_row+2)
            row_end = (last_row + 2)
            for r in range(3, row_end + 1):
                c = 2
                while c <= max_col:
                    cell = ws.cell(row=r, column=c)
                    val = cell.value
                    if not val:
                        c += 1
                        continue

                    # Expect format 'COURSE_CODE\nROOM' — extract course code
                    course_code = str(val).split('\n')[0].strip()
                    ctype = course_map.get(course_code, '')
                    if ctype == 'LAB' and c + 1 <= max_col:
                        next_val = ws.cell(row=r, column=c+1).value
                        if next_val and str(next_val).split('\n')[0].strip() == course_code:
                            # merge c and c+1
                            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
                            master = ws.cell(row=r, column=c)
                            master.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            # apply borders
                            for cc in (c, c+1):
                                ws.cell(row=r, column=cc).border = thin
                            c += 2
                            continue
                    c += 1

        wb.save(out_path)
        print(f"✓ Export successful!\n  Output: {out_path}\n")
    except Exception as e:
        print(f"✗ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()
if __name__ == "__main__":
    main()
