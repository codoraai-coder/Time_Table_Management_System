"""
Excel Exporter Service - NEW FORMAT
Exports timetable with:
- Columns: Time slots (08:00-09:00 ... 17:00-18:00)
- Rows: Days (Monday - Friday)
- Cell format: Subject code (stacked) Classroom
"""

import json
from datetime import datetime, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.database import SessionLocal
from app.models import TimetableVersion, Course, Faculty, Section, Room, Timeslot


class ExcelExporterServiceV2:
    """Export timetable to Excel with new format (time columns, days rows)."""
    
    DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    DAYS_SHORT = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    
    # All time slots from 8:00 to 18:00 (fixed)
    ALL_TIME_SLOTS = [
        "08:00-09:00", "09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00",
        "13:00-14:00", "14:00-15:00", "15:00-16:00", "16:00-17:00", "17:00-18:00"
    ]
    
    # Lunch break slots by shift
    LUNCH_BREAKS = {
        "SHIFT_8_4": "12:00-13:00",     # 8:00-16:00 shift has lunch at 12:00-13:00
        "SHIFT_10_6": "13:00-14:00"     # 10:00-18:00 shift has lunch at 13:00-14:00
    }
    
    def __init__(self, db: Session = None):
        self.db = db or SessionLocal()
    
    def export_timetable(self, version_id: int, output_path: str) -> bool:
        """
        Export timetable to Excel.
        
        Args:
            version_id: TimetableVersion ID to export
            output_path: Output file path
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Fetch timetable version
            version = self.db.query(TimetableVersion).filter_by(id=version_id).first()
            if not version:
                print(f"✗ Timetable version {version_id} not found")
                return False
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)
            
            # Get all sections
            sections = self.db.query(Section).order_by(Section.id).all()
            
            # Create sheet for each section
            for section in sections:
                self._create_section_sheet(wb, version, section)
            
            # Save workbook
            wb.save(output_path)
            print(f"[OK] Export successful! Output: {output_path}")
            return True
            
        except Exception as e:
            print(f"[-] Export failed: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _create_section_sheet(self, wb: Workbook, version: TimetableVersion, section: Section):
        """Create timetable sheet for a section, always showing all required slots and lunch break."""
        # Always use ALL_TIME_SLOTS for columns
        time_slots = self.ALL_TIME_SLOTS.copy()
        # Insert lunch break at correct slot for this section's shift
        shift_code = section.shift
        lunch_slot = self.LUNCH_BREAKS.get(shift_code)
        # Ensure lunch slot is present in time_slots (should be by default)
        # (If not, insert at correct position)
        if lunch_slot and lunch_slot not in time_slots:
            # Insert at correct position based on time
            lunch_index = 0
            for i, slot in enumerate(time_slots):
                if slot > lunch_slot:
                    lunch_index = i
                    break
            else:
                lunch_index = len(time_slots)
            time_slots.insert(lunch_index, lunch_slot)

        # Create sheet
        sheet_name = f"Section {section.name}"
        ws = wb.create_sheet(sheet_name)

        # Parse snapshot data
        snapshot_data = version.snapshot_data
        if isinstance(snapshot_data, str):
            snapshot_data = json.loads(snapshot_data)

        # Get section data from snapshot
        section_data = snapshot_data.get('sections', {}).get(section.code, {})

        # Set up borders and fonts
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        # 1. Title row
        ws['A1'] = f"Section {section.name} Timetable"
        ws['A1'].font = Font(bold=True, size=14, color="000000")
        ws.merge_cells(f'A1:{get_column_letter(len(time_slots) + 1)}1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # 2. Header row - Time slots
        ws['A2'] = "Day/Time"
        ws['A2'].font = header_font
        ws['A2'].fill = header_fill
        ws['A2'].border = thin_border
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[2].height = 30

        col = 2  # Start from column B
        for time_slot in time_slots:
            cell = ws.cell(row=2, column=col)
            cell.value = time_slot
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = 15
            col += 1

        # 3. Day rows
        row = 3
        for day in self.DAYS:
            # Day cell
            day_cell = ws.cell(row=row, column=1)
            day_cell.value = day
            day_cell.font = Font(bold=True)
            day_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            day_cell.border = thin_border
            day_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Time slot cells for this day
            col = 2
            day_assignments = section_data.get(day, [])

            for time_slot in time_slots:
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
                # If this is the lunch break slot, always leave empty
                if lunch_slot and time_slot == lunch_slot:
                    # If solver incorrectly placed an assignment at lunch, show it but warn.
                    assignment = self._find_assignment_by_time(day_assignments, time_slot)
                    if assignment:
                        cell.value = f"{assignment.get('course_code', '')}\n{assignment.get('room', '')}"
                        print(f"[WARN] Lunch slot occupied for section {section.code} on {day} at {time_slot}")
                    else:
                        cell.value = ""
                else:
                    # Find assignment for this time slot
                    assignment = self._find_assignment_by_time(day_assignments, time_slot)
                    if assignment:
                        course_code = assignment.get('course_code', '')
                        room_code = assignment.get('room', '')
                        cell.value = f"{course_code}\n{room_code}"
                ws.column_dimensions[get_column_letter(col)].width = 15
                ws.row_dimensions[row].height = 40
                col += 1
            row += 1

        # 4. Faculty section (below timetable)
        faculty_row = row + 2
        ws[f'A{faculty_row}'] = "Faculty Assignment"
        ws[f'A{faculty_row}'].font = Font(bold=True, size=11)
        faculty_row += 1
        
        # Collect unique courses for this section
        unique_assignments = {}
        for day_data in section_data.values():
            for assignment in day_data:
                course_code = assignment.get('course_code', '')
                if course_code and course_code not in unique_assignments:
                    unique_assignments[course_code] = assignment
        
        # Display faculty assignments
        for course_code, assignment in sorted(unique_assignments.items()):
            faculty_name = assignment.get('faculty', '')
            cell = ws[f'A{faculty_row}']
            cell.value = f"{course_code} | {faculty_name}"
            cell.font = Font(size=10)
            faculty_row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 18
    
    def _parse_shift(self, shift_code: str) -> tuple:
        """Parse shift code to start and end time."""
        shift_map = {
            "SHIFT_8_4": (time(8, 0), time(16, 0)),    # 8:00 to 16:00
            "SHIFT_10_6": (time(10, 0), time(18, 0)),  # 10:00 to 18:00
        }
        return shift_map.get(shift_code, (time(8, 0), time(16, 0)))
    
    def _generate_time_slots(self, start: time, end: time) -> list:
        """Generate hourly time slots between start and end time."""
        slots = []
        current = datetime.combine(datetime.today(), start)
        end_dt = datetime.combine(datetime.today(), end)
        
        while current < end_dt:
            next_hour = current + timedelta(hours=1)
            slot_str = f"{current.strftime('%H:%M')}-{next_hour.strftime('%H:%M')}"
            slots.append(slot_str)
            current = next_hour
        
        return slots
    
    def _find_assignment_by_time(self, day_assignments: list, time_slot: str) -> dict:
        """Find assignment for a specific time slot."""
        for assignment in day_assignments:
            assignment_time = assignment.get('time', '')
            # Normalize time format: "11:00 - 12:00" -> "11:00-12:00"
            normalized_assignment_time = assignment_time.replace(' - ', '-')
            if normalized_assignment_time == time_slot:
                return assignment
        return None
    
    def _get_assignments_for_section(self, snapshot_data: dict, section_id: int) -> list:
        """Extract assignments for a section from snapshot data."""
        assignments = []
        
        if 'assignments' not in snapshot_data:
            return assignments
        
        for assignment_data in snapshot_data['assignments']:
            if assignment_data.get('section_id') == section_id:
                # Get course, faculty, room details
                course = self.db.query(Course).filter_by(id=assignment_data.get('course_id')).first()
                faculty = self.db.query(Faculty).filter_by(id=assignment_data.get('faculty_id')).first()
                room = self.db.query(Room).filter_by(id=assignment_data.get('room_id')).first()
                timeslot = self.db.query(Timeslot).filter_by(id=assignment_data.get('timeslot_id')).first()
                
                if course and faculty and room and timeslot:
                    assignments.append({
                        'course_id': course.id,
                        'course_code': course.code,
                        'course_name': course.name,
                        'course_type': course.type,
                        'faculty_id': faculty.id,
                        'faculty_name': faculty.name,
                        'room_id': room.id,
                        'room_code': room.code,
                        'timeslot_id': timeslot.id,
                        'day_index': timeslot.day,
                        'start_time': timeslot.start_time,
                        'end_time': timeslot.end_time,
                    })
        
        return assignments
    
    def _find_assignment(self, assignments: list, day_idx: int, slot_idx: int, time_slots: list) -> dict:
        """Find assignment for a specific day and time slot."""
        
        # Parse target time from time slot string
        target_slot = time_slots[slot_idx]
        target_start = target_slot.split('-')[0]  # e.g., "10:00"
        
        for assignment in assignments:
            # Compare day
            if assignment.get('day_index') != day_idx:
                continue
            
            # Compare time (convert to HH:MM format)
            slot_time = assignment.get('start_time')
            if isinstance(slot_time, str):
                assignment_start = slot_time[:5]  # First 5 chars (HH:MM)
            else:
                assignment_start = slot_time.strftime('%H:%M')
            
            if assignment_start == target_start:
                return assignment
        
        return None
"""
Excel Exporter Service for Timetable Management System.

Exports TimetableVersion data to Excel format matching the required template:
- Grid layout: Rows = Time slots (8:00-18:00), Columns = Days (Mon-Fri)
- All sections in single sheet
- Lab merging: 2 consecutive rows for labs
- Cell format: CourseCode + Semester, Section
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from datetime import time
from typing import Dict, List, Optional, Tuple, Any
import json
from sqlalchemy.orm import Session
from sqlalchemy import select

from app.models import TimetableVersion, Course, Section


class ExcelExporterService:
    """Service to export timetables to Excel format."""
    
    def __init__(self, db: Session):
        self.db = db
        self.workbook = None
        self.sheet = None
        self.merged_cells = set()  # Track merged cells to avoid double-processing
    
    def export_timetable(self, version_id: int, output_path: str = "timetable_export.xlsx") -> str:
        """
        Export a timetable version to Excel format.
        
        Args:
            version_id: ID of TimetableVersion to export
            output_path: Output file path
            
        Returns:
            Path to created Excel file
        """
        # Fetch timetable version
        version = self.db.query(TimetableVersion).filter(
            TimetableVersion.id == version_id
        ).first()
        
        if not version:
            raise ValueError(f"TimetableVersion with ID {version_id} not found")
        
        # Initialize workbook
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Timetable"
        
        # Build and populate grid
        self._create_grid_structure()
        self._populate_assignments(version)
        self._apply_formatting()
        
        # Save file
        self.workbook.save(output_path)
        return output_path
    
    def _create_grid_structure(self):
        """Create the grid structure with time slots and days."""
        # Days of week
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        
        # Generate time slots (8:00 to 18:00)
        time_slots = self._generate_time_slots()
        
        # Headers
        self.sheet["A1"] = "Day/Time"
        for col_idx, day in enumerate(days, start=2):
            self.sheet.cell(row=1, column=col_idx, value=day)
        
        # Time slot labels in first column
        for row_idx, slot_time in enumerate(time_slots, start=2):
            self.sheet.cell(row=row_idx, column=1, value=slot_time)
        
        # Set column widths
        self.sheet.column_dimensions["A"].width = 15
        for col in range(2, 7):
            self.sheet.column_dimensions[get_column_letter(col)].width = 30
        
        # Set row height
        for row in range(2, len(time_slots) + 2):
            self.sheet.row_dimensions[row].height = 50
    
    def _generate_time_slots(self) -> List[str]:
        """Generate time slot labels from 8:00 to 18:00."""
        slots = []
        for hour in range(8, 18):
            slot = f"{hour}:00 - {hour + 1}:00"
            slots.append(slot)
        return slots
    
    def _get_timeslot_row(self, start_time: time) -> int:
        """
        Get the Excel row number for a given start time.
        
        Args:
            start_time: Start time of the slot
            
        Returns:
            Excel row number (1-based)
        """
        # Row 1 is header, Row 2 starts at 8:00
        # So 8:00 = row 2, 9:00 = row 3, etc.
        return start_time.hour - 8 + 2
    
    def _get_day_column(self, day_index: int) -> int:
        """
        Get the Excel column number for a day.
        
        Args:
            day_index: Day of week (0=Monday, 4=Friday)
            
        Returns:
            Excel column number (1-based)
        """
        # Column 1 is time labels, Column 2 is Monday, etc.
        return day_index + 2
    
    def _populate_assignments(self, version: TimetableVersion):
        """
        Populate the grid with assignments from the timetable version snapshot.
        
        Args:
            version: TimetableVersion object containing snapshot_data
        """
        snapshot_data = version.snapshot_data
        if not snapshot_data or "sections" not in snapshot_data:
            return
        
        sections_data = snapshot_data.get("sections", {})
        
        # Process each section in snapshot
        for section_name, days_data in sections_data.items():
            # days_data format: {day: [assignments]}
            for day_name, assignments_list in days_data.items():
                if not assignments_list:
                    continue
                
                day_index = self._day_name_to_index(day_name)
                
                for assignment_data in assignments_list:
                    self._populate_assignment_from_data(
                        assignment_data, section_name, day_index
                    )
    
    def _day_name_to_index(self, day_name: str) -> int:
        """Convert day name to index (0=Monday, 4=Friday)."""
        days = {
            "Monday": 0,
            "Tuesday": 1,
            "Wednesday": 2,
            "Thursday": 3,
            "Friday": 4
        }
        return days.get(day_name, 0)
    
    def _populate_assignment_from_data(self, assignment_data: Dict[str, Any], section_name: str, day_index: int):
        """
        Populate a single assignment from snapshot data.
        
        Args:
            assignment_data: Assignment data from snapshot
            section_name: Section name
            day_index: Day of week (0-4)
        """
        # Extract data from snapshot
        time_str = assignment_data.get("time", "8:00 - 9:00")
        course_code = assignment_data.get("course_code", "UNKNOWN")
        course_name = assignment_data.get("course", "UNKNOWN")
        
        # Parse time to get start hour
        start_hour = int(time_str.split(":")[0])
        start_time = time(start_hour, 0)
        
        # Determine grid position
        row = self._get_timeslot_row(start_time)
        col = self._get_day_column(day_index)
        
        # Skip if already merged
        cell_key = (row, col)
        if cell_key in self.merged_cells:
            return
        
        # Check if this is a lab (labs have 2-hour slots)
        is_lab = "LAB" in course_code.upper()
        
        # Format cell content
        cell_content = f"{course_code}\n{section_name}"
        
        # Handle lab merging (2 consecutive rows)
        if is_lab:
            next_row = row + 1
            if next_row <= 26:  # Max row for 18:00
                # Merge cells
                merge_range = f"{get_column_letter(col)}{row}:{get_column_letter(col)}{next_row}"
                self.sheet.merge_cells(merge_range)
                
                # Mark both cells as merged
                self.merged_cells.add((row, col))
                self.merged_cells.add((next_row, col))
        
        # Write content
        cell = self.sheet.cell(row=row, column=col, value=cell_content)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def _apply_formatting(self):
        """Apply formatting to the entire sheet."""
        # Define border style
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Apply to all cells with data
        for row in self.sheet.iter_rows(min_row=1, max_row=26, min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Bold header row
        for cell in self.sheet[1]:
            if cell.value:
                cell.font = Font(bold=True)
        
        # Bold time column
        for row in range(2, 27):
            self.sheet.cell(row=row, column=1).font = Font(bold=True)


# Example usage function
def export_example(db: Session, version_id: int = 1):
    """Export a timetable version to Excel."""
    exporter = ExcelExporterService(db)
    try:
        output_file = exporter.export_timetable(version_id, "timetable_export.xlsx")
        print(f"✓ Timetable exported successfully to: {output_file}")
        return output_file
    except Exception as e:
        print(f"✗ Error exporting timetable: {e}")
        raise
